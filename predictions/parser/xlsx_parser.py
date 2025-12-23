"""Parser for family prediction game xlsx files."""

from dataclasses import dataclass
from datetime import date
from enum import Enum
from pathlib import Path
from typing import Any, Optional

import openpyxl
import pandas as pd


class SheetStructure(Enum):
    """Identifies the xlsx file structure format."""

    LEGACY = "legacy"  # 2022-2024: Uses 'Main' sheet
    CURRENT = "current"  # 2025+: Uses 'Statements' sheet


@dataclass
class Statement:
    """Represents a prediction statement/question."""

    id: str
    year: int
    text: str
    category: str
    proposer: str


@dataclass
class Prediction:
    """Represents a participant's probability assignment to a statement."""

    statement_id: str
    participant: str
    probability: float


@dataclass
class Outcome:
    """Represents the resolved outcome of a statement."""

    statement_id: str
    outcome: Optional[bool]
    date: Optional[date]


@dataclass
class GameData:
    """Container for all game data from a year."""

    statements: list[Statement]
    predictions: list[Prediction]
    outcomes: list[Outcome]


def validate_probability(value: Any) -> float:
    """
    Validate and convert a probability value.

    Args:
        value: The probability value to validate

    Returns:
        The probability as a float in range [0.0, 1.0]

    Raises:
        ValueError: If the value is not a valid probability
    """
    # Handle None
    if value is None:
        raise ValueError("Probability cannot be None")

    # Try to convert to float
    try:
        prob = float(value)
    except (TypeError, ValueError) as e:
        raise ValueError(f"Cannot convert {value!r} to probability: {e}")

    # Check for NaN
    if prob != prob:  # NaN check
        raise ValueError("Probability cannot be NaN")

    # Check range
    if not (0.0 <= prob <= 1.0):
        raise ValueError(f"Probability must be in range [0.0, 1.0], got: {prob}")

    return prob


def sanitize_probability(value: Any) -> float:
    """
    Sanitize and convert a probability value, using 0.5 as fallback for invalid values.

    Handles:
    - NaN -> 0.5
    - None -> 0.5
    - Values < 0.0 -> 0.5
    - Values > 1.0 -> 0.5
    - Invalid conversions -> 0.5

    Args:
        value: The probability value to sanitize

    Returns:
        The probability as a float in range [0.0, 1.0], or 0.5 if invalid
    """
    # Handle None
    if value is None:
        return 0.5

    # Try to convert to float
    try:
        prob = float(value)
    except (TypeError, ValueError):
        return 0.5

    # Check for NaN
    if prob != prob:  # NaN check
        return 0.5

    # Check range
    if not (0.0 <= prob <= 1.0):
        return 0.5

    return prob


def parse_outcome(value: Any) -> Optional[bool]:
    """
    Parse outcome value to Optional[bool].

    Handles different formats across years:
    - 0 or 0.0 -> False
    - 1 or 1.0 -> True
    - NaN, None, "-", or empty string -> None

    Args:
        value: The outcome value from the xlsx file

    Returns:
        True for positive outcomes, False for negative, None for unresolved

    Raises:
        ValueError: If the value cannot be parsed
    """
    # Handle None
    if value is None:
        return None

    # Handle string values
    if isinstance(value, str):
        value_stripped = value.strip()
        if value_stripped == "" or value_stripped == "-":
            return None
        if value_stripped in ("0", "0.0"):
            return False
        if value_stripped in ("1", "1.0"):
            return True
        raise ValueError(f"Unable to parse outcome value: {value!r}")

    # Handle numeric values
    if isinstance(value, (int, float)):
        # Check for NaN
        if value != value:  # NaN check
            return None
        if value == 0 or value == 0.0:
            return False
        if value == 1 or value == 1.0:
            return True
        raise ValueError(f"Numeric outcome must be 0 or 1, got: {value}")

    raise ValueError(f"Unexpected outcome type: {type(value).__name__}")


def extract_first_proposer(name: str) -> str:
    """
    Extract the first name from a multi-person proposer string.

    Handles formats like:
    - "james and chris" -> "james"
    - "iain, christine and andrew" -> "iain"
    - "alice & bob" -> "alice"

    Args:
        name: The proposer name(s) to parse

    Returns:
        The first proposer name
    """
    import re

    # Split on common separators: ", " or " and " or " & "
    # Use regex to split on any of these patterns
    parts = re.split(r"\s*,\s*|\s+and\s+|\s*&\s*", name, flags=re.IGNORECASE)

    # Return the first non-empty part, stripped of whitespace
    for part in parts:
        stripped = part.strip()
        if stripped:
            return stripped

    # Fallback to original if no parts found
    return name.strip()


def normalize_participant_name(name: str) -> str:
    """
    Normalize participant names to handle variations.

    Handles:
    - Bruce/Boof -> Bruce
    - Gaël/Gael/Gäel -> Gael
    - Christine/Chris -> Christine

    Args:
        name: The participant name to normalize

    Returns:
        The normalized participant name
    """
    # Strip whitespace and convert to consistent case for comparison
    normalized = name.strip()

    # Create a mapping of variants to canonical names
    name_mapping = {
        # Bruce variants
        "boof": "Bruce",
        "bruce": "Bruce",
        # Gael variants (handle accent variations)
        "gael": "Gael",
        "gaël": "Gael",
        "gäel": "Gael",
        # Christine variants
        "chris": "Christine",
        "christine": "Christine",
    }

    # Look up the canonical name (case-insensitive)
    canonical = name_mapping.get(normalized.lower())

    return canonical if canonical else normalized


def detect_structure(xlsx_path: Path) -> SheetStructure:
    """
    Detect which xlsx structure format is used.

    Args:
        xlsx_path: Path to the xlsx file

    Returns:
        SheetStructure indicating the format type

    Raises:
        ValueError: If neither Main/Input nor Statements sheet is found
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    if "Statements" in sheet_names:
        return SheetStructure.CURRENT
    elif "Main" in sheet_names or "Input" in sheet_names:
        return SheetStructure.LEGACY
    else:
        raise ValueError(
            f"Unable to detect structure for {xlsx_path}: "
            f"Neither 'Main'/'Input' nor 'Statements' sheet found. "
            f"Available sheets: {sheet_names}"
        )


def extract_statements_2022_2024(filepath: Path, year: int) -> list[Statement]:
    """
    Extract statements from 2022-2024 format Excel files.

    Reads the 'Main' or 'Input' sheet and extracts statement data.

    Args:
        filepath: Path to the Excel file
        year: The year of the predictions

    Returns:
        List of Statement objects

    Raises:
        ValueError: If required fields are missing or IDs are not unique
    """
    # Determine which sheet name to use
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    if "Main" in sheet_names:
        sheet_name = "Main"
    elif "Input" in sheet_names:
        sheet_name = "Input"
    else:
        raise ValueError(f"Neither 'Main' nor 'Input' sheet found in {filepath}")

    # Read the sheet
    df = pd.read_excel(filepath, sheet_name=sheet_name)

    # Validate required columns exist
    required_columns = ["ID", "Prediction", "Category", "Proposer"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(
            f"Missing required columns in Main sheet: {missing_columns}"
        )

    # Drop rows where ID is NaN (these are empty rows)
    df = df.dropna(subset=["ID"])

    # Check for unique IDs
    if df["ID"].duplicated().any():
        duplicates = df[df["ID"].duplicated(keep=False)]["ID"].tolist()
        raise ValueError(f"Duplicate statement IDs found: {duplicates}")

    # Extract statements
    statements = []
    for _, row in df.iterrows():
        # Convert ID to string
        statement_id = str(int(row["ID"]))

        # Extract first name if multiple proposers, then normalize
        proposer = normalize_participant_name(extract_first_proposer(str(row["Proposer"])))

        # Create Statement object with globally unique ID (year-id)
        global_id = f"{year}-{statement_id}"
        statement = Statement(
            id=global_id,
            year=year,
            text=str(row["Prediction"]).strip(),
            category=str(row["Category"]).strip(),
            proposer=proposer,
        )
        statements.append(statement)

    return statements


def extract_statements_2025(filepath: Path, year: int) -> list[Statement]:
    """
    Extract statements from 2025 format Excel files.

    Reads the 'Statements' sheet and extracts statement data.

    Args:
        filepath: Path to the Excel file
        year: The year of the predictions

    Returns:
        List of Statement objects

    Raises:
        ValueError: If required fields are missing or IDs are not unique
    """
    # Read the Statements sheet
    df = pd.read_excel(filepath, sheet_name="Statements")

    # Determine which proposer column to use (Name or Author)
    if "Name" in df.columns:
        proposer_col = "Name"
    elif "Author" in df.columns:
        proposer_col = "Author"
    else:
        raise ValueError(
            "Missing proposer column in Statements sheet (expected 'Name' or 'Author')"
        )

    # Validate required columns exist
    required_columns = ["Number", proposer_col, "Prediction", "Category"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(
            f"Missing required columns in Statements sheet: {missing_columns}"
        )

    # Drop rows where Number is NaN (these are empty rows)
    df = df.dropna(subset=["Number"])

    # Check for unique IDs
    if df["Number"].duplicated().any():
        duplicates = df[df["Number"].duplicated(keep=False)]["Number"].tolist()
        raise ValueError(f"Duplicate statement IDs found: {duplicates}")

    # Extract statements
    statements = []
    for _, row in df.iterrows():
        # Convert Number to string (this is the ID)
        statement_id = str(int(row["Number"]))

        # Extract first name if multiple proposers, then normalize
        proposer = normalize_participant_name(extract_first_proposer(str(row[proposer_col])))

        # Create Statement object with globally unique ID (year-id)
        global_id = f"{year}-{statement_id}"
        statement = Statement(
            id=global_id,
            year=year,
            text=str(row["Prediction"]).strip(),
            category=str(row["Category"]).strip(),
            proposer=proposer,
        )
        statements.append(statement)

    return statements


def extract_predictions_2022_2024(
    filepath: Path, statements: list[Statement]
) -> list[Prediction]:
    """
    Extract predictions from 2022-2024 format Excel files.

    Reads participant sheets and extracts prediction data.

    Args:
        filepath: Path to the Excel file
        statements: List of Statement objects to validate against

    Returns:
        List of Prediction objects

    Raises:
        ValueError: If validation fails
    """
    # Create a set of valid statement IDs for validation
    valid_ids = {stmt.id for stmt in statements}

    # Get all sheet names
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    # Exclude non-participant sheets
    excluded_sheets = {"Main", "Summary", "Scores"}
    participant_sheets = [
        name for name in sheet_names if name not in excluded_sheets
    ]

    predictions = []

    for sheet_name in participant_sheets:
        # Normalize participant name from sheet name
        participant = normalize_participant_name(sheet_name)

        # Skip ChatGPT participant
        if participant.lower() == "chatgpt":
            continue

        # Read the participant sheet
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
        except Exception as e:
            raise ValueError(f"Error reading sheet {sheet_name}: {e}")

        # Check for required columns
        if "ID" not in df.columns:
            continue

        # Determine which probability column to use
        if "Probability (0.0 to 1.0)" in df.columns:
            prob_col = "Probability (0.0 to 1.0)"
        elif "Probability" in df.columns:
            prob_col = "Probability"
        else:
            # Skip sheets that don't have a probability column
            continue

        # Drop rows where ID is NaN
        df = df.dropna(subset=["ID"])

        # Extract predictions
        for _, row in df.iterrows():
            statement_id = str(int(row["ID"]))

            # Create globally unique statement ID by finding the year
            # The statement_id here is the local ID, we need to find which year's statements to match
            global_id = None
            for stmt in statements:
                # Extract local ID from global ID (format: "year-id")
                if "-" in stmt.id:
                    local_id = stmt.id.split("-", 1)[1]
                    if local_id == statement_id:
                        global_id = stmt.id
                        break

            if global_id is None:
                raise ValueError(
                    f"Invalid statement ID {statement_id} in sheet {sheet_name}"
                )

            # Get and sanitize probability (converts invalid values to 0.5)
            prob_value = row[prob_col]
            probability = sanitize_probability(prob_value)

            prediction = Prediction(
                statement_id=global_id,
                participant=participant,
                probability=probability,
            )
            predictions.append(prediction)

    return predictions


def normalize_text_for_matching(text: str) -> str:
    """
    Normalize text for matching by stripping, normalizing whitespace, and removing trailing punctuation.

    Args:
        text: The text to normalize

    Returns:
        Normalized text
    """
    # Strip whitespace and convert to lowercase for matching
    normalized = text.strip().lower()
    # Normalize internal whitespace (replace multiple spaces with single space)
    normalized = " ".join(normalized.split())
    # Remove trailing punctuation
    while normalized and normalized[-1] in ".!?,;:":
        normalized = normalized[:-1]
    return normalized


def extract_predictions_2025(
    filepath: Path, statements: list[Statement]
) -> list[Prediction]:
    """
    Extract predictions from 2025 format Excel files.

    Reads participant sheets and matches predictions by text.

    Args:
        filepath: Path to the Excel file
        statements: List of Statement objects to match against

    Returns:
        List of Prediction objects

    Raises:
        ValueError: If validation fails
    """
    # Create a mapping from normalized prediction text to statement
    # (we need the full statement to get the globally unique ID)
    text_to_stmt = {}
    for stmt in statements:
        normalized = normalize_text_for_matching(stmt.text)
        text_to_stmt[normalized] = stmt

    # Get all sheet names
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    # Exclude non-participant sheets
    excluded_sheets = {"Statements", "Summary"}
    participant_sheets = [
        name for name in sheet_names if name not in excluded_sheets
    ]

    predictions = []

    for sheet_name in participant_sheets:
        # Normalize participant name from sheet name
        participant = normalize_participant_name(sheet_name)

        # Skip ChatGPT participant
        if participant.lower() == "chatgpt":
            continue

        # Read the participant sheet
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
        except Exception as e:
            raise ValueError(f"Error reading sheet {sheet_name}: {e}")

        # Check for required columns
        if "Prediction" not in df.columns or "Probability" not in df.columns:
            # Skip sheets that don't have the expected format
            continue

        # Drop rows where Prediction is NaN
        df = df.dropna(subset=["Prediction"])

        # Extract predictions
        for idx, row in df.iterrows():
            prediction_text = str(row["Prediction"]).strip()

            # Skip Excel error references
            if prediction_text.startswith("#REF!") or prediction_text == "#REF!":
                continue

            # Match to statement by normalized text
            normalized_text = normalize_text_for_matching(prediction_text)
            stmt = text_to_stmt.get(normalized_text)
            if stmt is None:
                raise ValueError(
                    f"Cannot match prediction text in sheet {sheet_name} at row {idx}: "
                    f"{prediction_text[:100]}..."
                )

            # Get and sanitize probability (converts invalid values to 0.5)
            prob_value = row["Probability"]
            probability = sanitize_probability(prob_value)

            prediction = Prediction(
                statement_id=stmt.id,  # Use globally unique ID
                participant=participant,
                probability=probability,
            )
            predictions.append(prediction)

    return predictions


def extract_outcomes(filepath: Path, statements: list[Statement]) -> list[Outcome]:
    """
    Extract outcomes from Excel files (both formats).

    Args:
        filepath: Path to the Excel file
        statements: List of Statement objects to match against

    Returns:
        List of Outcome objects
    """
    structure = detect_structure(filepath)

    if structure == SheetStructure.LEGACY:
        # 2022-2024: Read from Main or Input sheet
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        if "Main" in sheet_names:
            sheet_name = "Main"
        elif "Input" in sheet_names:
            sheet_name = "Input"
        else:
            raise ValueError(f"Neither 'Main' nor 'Input' sheet found in {filepath}")

        df = pd.read_excel(filepath, sheet_name=sheet_name)

        # Check which outcome column name is used
        if "Outcome (1 = true, 0 = false)" in df.columns:
            outcome_col = "Outcome (1 = true, 0 = false)"
        elif "Outcome" in df.columns:
            outcome_col = "Outcome"
        else:
            outcome_col = None

        # Special case for 2022: outcomes are in Summary sheet, not Main
        # Check if Main sheet outcome column is empty/all NaN
        if sheet_name == "Main" and (
            outcome_col is None or df[outcome_col].isna().all()
        ):
            if "Summary" in sheet_names:
                df = pd.read_excel(filepath, sheet_name="Summary")
                # 2022 Summary uses different column name
                if "Outcome (0 = false, 1 = true)" in df.columns:
                    outcome_col = "Outcome (0 = false, 1 = true)"
                elif "Outcome (1 = true, 0 = false)" in df.columns:
                    outcome_col = "Outcome (1 = true, 0 = false)"
                else:
                    outcome_col = "Outcome"
    else:
        # 2025: Read from Statements sheet
        df = pd.read_excel(filepath, sheet_name="Statements")
        # Check which outcome column name is used
        if "Outcome (1 = true, 0 = false)" in df.columns:
            outcome_col = "Outcome (1 = true, 0 = false)"
        else:
            outcome_col = "Outcome"

    # Create a mapping from statement ID to outcome
    outcomes = []

    for stmt in statements:
        # Extract local ID from global ID (format: "year-id")
        if "-" in stmt.id:
            local_id = stmt.id.split("-", 1)[1]
        else:
            local_id = stmt.id

        # Find the row with this statement ID
        if structure == SheetStructure.LEGACY:
            row_match = df[df["ID"] == int(local_id)]
        else:
            row_match = df[df["Number"] == int(local_id)]

        if len(row_match) == 0:
            # No outcome data for this statement
            outcome_value = None
        else:
            outcome_value = row_match.iloc[0][outcome_col]

        # Parse the outcome (use globally unique ID)
        outcome = Outcome(
            statement_id=stmt.id,
            outcome=parse_outcome(outcome_value),
            date=None,  # Date parsing not yet implemented
        )
        outcomes.append(outcome)

    return outcomes


def parse_xlsx_file(filepath: Path) -> GameData:
    """
    Parse an Excel file and extract all game data.

    Auto-detects the file structure and routes to appropriate extraction functions.

    Args:
        filepath: Path to the Excel file

    Returns:
        GameData object containing all statements, predictions, and outcomes

    Raises:
        ValueError: If the file cannot be parsed or validation fails
    """
    import re

    # Convert to Path if string
    filepath = Path(filepath)

    # Extract year from filename (e.g., "2022.xlsx" or "test_2022.xlsx" -> 2022)
    # Look for a 4-digit year (2000-2099) in the filename
    match = re.search(r'20\d{2}', filepath.stem)
    if match:
        year = int(match.group())
    else:
        raise ValueError(f"Cannot extract year from filename: {filepath.name}")

    # Detect structure
    structure = detect_structure(filepath)

    # Extract statements
    if structure == SheetStructure.LEGACY:
        statements = extract_statements_2022_2024(filepath, year)
    else:
        statements = extract_statements_2025(filepath, year)

    # Extract predictions
    if structure == SheetStructure.LEGACY:
        predictions = extract_predictions_2022_2024(filepath, statements)
    else:
        predictions = extract_predictions_2025(filepath, statements)

    # Extract outcomes
    outcomes = extract_outcomes(filepath, statements)

    # Create and return GameData
    return GameData(
        statements=statements,
        predictions=predictions,
        outcomes=outcomes,
    )


def parse_all_years(input_dir: Path) -> GameData:
    """
    Parse all Excel files in a directory and aggregate into a single GameData object.

    Args:
        input_dir: Directory containing year xlsx files

    Returns:
        GameData object containing all statements, predictions, and outcomes from all years

    Raises:
        ValueError: If validation fails (e.g., duplicate statement IDs within same year)
    """
    import re

    # Convert to Path if string
    input_dir = Path(input_dir)

    # Find all .xlsx files
    xlsx_files = sorted(input_dir.glob("*.xlsx"))

    if not xlsx_files:
        raise ValueError(f"No .xlsx files found in {input_dir}")

    # Parse each file and collect data
    all_statements = []
    all_predictions = []
    all_outcomes = []

    # Track statement IDs by year to validate uniqueness within each year
    statements_by_year = {}

    for xlsx_file in xlsx_files:
        try:
            game_data = parse_xlsx_file(xlsx_file)

            # Extract year from filename using regex (handles test_2022.xlsx etc)
            match = re.search(r'20\d{2}', xlsx_file.stem)
            if match:
                year = int(match.group())
            else:
                # Fall back to using year from first statement
                year = game_data.statements[0].year if game_data.statements else None

            if year:
                if year not in statements_by_year:
                    statements_by_year[year] = set()

                # Validate no duplicate statement IDs within same year
                for stmt in game_data.statements:
                    if stmt.id in statements_by_year[year]:
                        raise ValueError(
                            f"Duplicate statement ID {stmt.id} found in year {year}"
                        )
                    statements_by_year[year].add(stmt.id)

            # Aggregate data
            all_statements.extend(game_data.statements)
            all_predictions.extend(game_data.predictions)
            all_outcomes.extend(game_data.outcomes)

        except Exception as e:
            # Re-raise with file context
            raise ValueError(f"Error parsing {xlsx_file.name}: {e}") from e

    # Create and return aggregated GameData
    return GameData(
        statements=all_statements,
        predictions=all_predictions,
        outcomes=all_outcomes,
    )


def export_to_csv(game_data: GameData, output_path: Path) -> None:
    """
    Export GameData to CSV format.

    Creates a denormalized CSV where each row is a statement with its outcome
    and all participants' predictions as columns.

    Args:
        game_data: The GameData object to export
        output_path: Path where the CSV file should be written

    Raises:
        IOError: If the file cannot be written
    """
    import csv

    # Create indices for quick lookups
    predictions_by_stmt = {}
    for pred in game_data.predictions:
        if pred.statement_id not in predictions_by_stmt:
            predictions_by_stmt[pred.statement_id] = {}
        predictions_by_stmt[pred.statement_id][pred.participant] = pred.probability

    outcomes_by_stmt = {outcome.statement_id: outcome for outcome in game_data.outcomes}

    # Get all unique participants (sorted for consistent column order)
    participants = sorted(
        {pred.participant for pred in game_data.predictions}
    )

    # Define CSV columns
    base_columns = ["id", "year", "text", "category", "proposer", "outcome", "outcome_date"]
    columns = base_columns + participants

    # Prepare rows
    rows = []
    for stmt in game_data.statements:
        row = {
            "id": stmt.id,
            "year": stmt.year,
            "text": stmt.text,
            "category": stmt.category,
            "proposer": stmt.proposer,
        }

        # Add outcome
        outcome = outcomes_by_stmt.get(stmt.id)
        if outcome:
            row["outcome"] = outcome.outcome
            row["outcome_date"] = outcome.date.isoformat() if outcome.date else ""
        else:
            row["outcome"] = ""
            row["outcome_date"] = ""

        # Add predictions for each participant
        stmt_predictions = predictions_by_stmt.get(stmt.id, {})
        for participant in participants:
            row[participant] = stmt_predictions.get(participant, "")

        rows.append(row)

    # Write to CSV
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def export_to_arrow(game_data: GameData, output_path: Path) -> None:
    """
    Export GameData to Apache Arrow format.

    Creates a denormalized Arrow file where each row is a statement with its outcome
    and all participants' predictions as columns. Compatible with arquero library.

    Args:
        game_data: The GameData object to export
        output_path: Path where the Arrow file should be written

    Raises:
        IOError: If the file cannot be written
    """
    import pyarrow as pa
    import pyarrow.feather as feather

    # Create indices for quick lookups
    predictions_by_stmt = {}
    for pred in game_data.predictions:
        if pred.statement_id not in predictions_by_stmt:
            predictions_by_stmt[pred.statement_id] = {}
        predictions_by_stmt[pred.statement_id][pred.participant] = pred.probability

    outcomes_by_stmt = {outcome.statement_id: outcome for outcome in game_data.outcomes}

    # Get all unique participants (sorted for consistent column order)
    participants = sorted(
        {pred.participant for pred in game_data.predictions}
    )

    # Prepare data as lists for each column
    data = {
        "id": [],
        "year": [],
        "text": [],
        "category": [],
        "proposer": [],
        "outcome": [],
        "outcome_date": [],
    }
    for participant in participants:
        data[participant] = []

    # Build rows
    for stmt in game_data.statements:
        data["id"].append(stmt.id)
        data["year"].append(stmt.year)
        data["text"].append(stmt.text)
        data["category"].append(stmt.category)
        data["proposer"].append(stmt.proposer)

        # Add outcome
        outcome = outcomes_by_stmt.get(stmt.id)
        if outcome and outcome.outcome is not None:
            data["outcome"].append(outcome.outcome)
        else:
            data["outcome"].append(None)

        if outcome and outcome.date:
            data["outcome_date"].append(outcome.date.isoformat())
        else:
            data["outcome_date"].append(None)

        # Add predictions for each participant
        stmt_predictions = predictions_by_stmt.get(stmt.id, {})
        for participant in participants:
            prob = stmt_predictions.get(participant)
            data[participant].append(prob if prob is not None else None)

    # Convert to pandas DataFrame for easier Arrow conversion
    df = pd.DataFrame(data)

    # Ensure output directory exists
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Write to Arrow format using Feather V2 (which is IPC format)
    # Disable compression for browser compatibility - the JS Arrow library
    # used by arquero doesn't support LZ4/ZSTD record batch decompression
    feather.write_feather(df, output_path, compression='uncompressed')
